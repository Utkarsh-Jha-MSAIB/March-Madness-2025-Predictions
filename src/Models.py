import pandas as pd
from sklearn.model_selection import train_test_split
import matplotlib.pyplot as plt
from math import sin, cos, sqrt, atan2, radians
import numpy as np
import random
from sklearn.model_selection import train_test_split
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import (auc, classification_report, roc_auc_score, accuracy_score,
                             f1_score, log_loss, roc_curve, confusion_matrix, precision_score, recall_score)
import statsmodels.formula.api as smf
from sklearn.decomposition import PCA
from sklearn.preprocessing import StandardScaler
from sklearn.ensemble import RandomForestClassifier
from sklearn.neural_network import MLPClassifier as MLP
from sklearn.inspection import permutation_importance
import xgboost as xgb
import ast
import os
from tqdm import tqdm
from openpyxl.styles import PatternFill
from openpyxl import load_workbook




def prepare_train_test_split(df: pd.DataFrame, split_type: str = 'yearly', 
                              split_year: int = 2010, train_size: float = 0.7, 
                              random_state: int = 42) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Splits the dataset into training and testing sets using either a fixed year or random split.
    Also generates a groupby summary table by 'team1_win'.

    Parameters:
    ----------
    df : pd.DataFrame
        The full input DataFrame to be split.

    split_type : str
        'yearly' (default) for a fixed year-based split, or 'random' for random train-test split.

    split_year : int
        The year to use as test set when split_type='yearly'. Default is 2010.

    train_size : float
        Proportion of data to use for training when split_type='random'. Default is 0.7.

    random_state : int
        Random seed for reproducibility (used only for random split).

    Returns:
    -------
    train : pd.DataFrame
        Training dataset

    test : pd.DataFrame
        Testing dataset

    summary_table : pd.DataFrame
        Group-level summary of selected features by 'team1_win'
    """

    if split_type == 'yearly':
        train = df.query("year != @split_year").reset_index(drop=True)
        test = df.query("year == @split_year").reset_index(drop=True)
    elif split_type == 'random':
        train, test = train_test_split(df, train_size=train_size, random_state=random_state)
        train = train.reset_index(drop=True)
        test = test.reset_index(drop=True)
    else:
        raise ValueError("split_type must be 'yearly' or 'random'")

    # Initial groupby summary
    summary_table = train.groupby('team1_win')[[
        'point_diff', 'team1_score', 'team2_score', 'PC1', 'TurnoverMargin'
    ]].mean().round(4).reset_index()

    return train, test, summary_table


# Defined function for building AUC Curve

def plot_roc(fpr, tpr, roc_auc):
    """Plots the ROC curve for the win probability model along with
    the AUC.
    """
    fig, ax = plt.subplots()
    ax.set(title='Receiver Operating Characteristic',
           xlim=[0, 1], ylim=[0, 1], xlabel='False Positive Rate',
           ylabel='True Positive Rate')
    ax.plot(fpr, tpr, 'b', label='AUC = %0.2f' % roc_auc)
    ax.plot([0, 1], [0, 1], 'k--')
    ax.legend(loc='lower right')


#Logistic model function

def run_logistic_model(
    fix_train: pd.DataFrame,
    df: pd.DataFrame, 
    ind_var: list[str], 
    dep_var: str = 'team1_win',
    data_label: str = 'train',
):
    """
    Trains and evaluates a logistic regression model on the specified dataset (train/test),
    prints performance metrics, plots ROC curve, and visualizes feature importance.

    Parameters:
    ----------
    df : pd.DataFrame
        DataFrame to train/test on.

    ind_var : list of str
        Independent variable names (features).

    dep_var : str
        Dependent variable (target). Default is 'team1_win'.

    data_label : str
        Label for logging purposes ('train' or 'test').

    Returns:
    -------
    dict
        Dictionary with performance metrics and feature importance indices.
    """
    # Logistic Regression
    logit = LogisticRegression()
    logit.fit(fix_train[ind_var], fix_train[dep_var])

    # Statsmodels summary (optional/logging only)
    ind_var_add = ' + '.join(ind_var)
    smf.logit(f'{dep_var} ~ {ind_var_add}', fix_train).fit(disp=False)

    # Predict probabilities and classes
    df = df.copy()
    df['prob'] = logit.predict_proba(df[ind_var])[:, 1]
    df['prediction'] = logit.predict(df[ind_var])

    # Evaluation
    print(f'\n📊 {data_label.upper()} SET EVALUATION')
    print('\nConfusion Matrix:')
    print(confusion_matrix(df[dep_var], df['prediction']))

    accuracy = accuracy_score(df[dep_var], df['prediction'])
    precision = precision_score(df[dep_var], df['prediction'])
    recall = recall_score(df[dep_var], df['prediction'])
    f1 = f1_score(df[dep_var], df['prediction'])
    ll = log_loss(df[dep_var].values, df['prob'].values, labels=[0, 1])

    print(f'\nAccuracy: {accuracy:.2%}')
    print(f'Precision: {precision:.2%}')
    print(f'Recall: {recall:.2%}')
    print(f'F1 Score: {f1:.2%}')
    print(f'Log Loss: {ll:.2%}')

    # ROC Curve
    fpr, tpr, _ = roc_curve(df[dep_var], df['prob'])
    roc_auc = auc(fpr, tpr)
    plot_roc(fpr, tpr, roc_auc)

    # Feature Importance
    feature_importance = abs(logit.coef_[0])
    feature_importance = 100 * (feature_importance / feature_importance.max())
    sorted_idx = np.argsort(feature_importance)[-10:]  # Top 10

    pos = np.arange(sorted_idx.shape[0]) + 0.5
    plt.figure()
    plt.barh(pos, feature_importance[sorted_idx], align='center', color='#1c60ae')
    plt.yticks(pos, np.array(ind_var)[sorted_idx], fontsize=8)
    plt.xlabel('Relative Feature Importance')
    plt.title(f'Feature Importance ({data_label})')
    plt.tight_layout()
    plt.show()

    return {
        'accuracy': accuracy,
        'precision': precision,
        'recall': recall,
        'f1': f1,
        'log_loss': ll,
        'roc_auc': roc_auc,
        'top_features_idx': sorted_idx
    }


#Random Forest Model

def run_random_forest_model(
    fix_train: pd.DataFrame,
    df: pd.DataFrame,
    ind_var: list[str],
    dep_var: str = 'team1_win',
    data_label: str = 'train',
    n_estimators: int = 100,
    max_depth: int = 7,
    min_samples_leaf: int = 2,
    min_samples_split: int = 5,
) -> dict:
    """
    Trains and evaluates a Random Forest classifier on the given dataset and prints performance metrics.

    Parameters:
    ----------
    df : pd.DataFrame
        Dataset to run the model on (train or test).

    ind_var : list of str
        Independent variable names.

    dep_var : str
        Target variable name. Default is 'team1_win'.

    data_label : str
        Label for display (e.g., 'train', 'test').

    n_estimators : int
        Number of trees in the forest.

    max_depth : int
        Max depth of the trees.

    min_samples_leaf : int
        Minimum number of samples required at a leaf node.

    min_samples_split : int
        Minimum number of samples required to split an internal node.

    Returns:
    -------
    dict
        Dictionary of model performance metrics and top feature indices.
    """
    # Initialize model
    rf = RandomForestClassifier(
        n_estimators=n_estimators,
        max_depth=max_depth,
        min_samples_leaf=min_samples_leaf,
        min_samples_split=min_samples_split,
        random_state=42
    )
    
    # Train the model
    rf.fit(fix_train[ind_var], fix_train[dep_var])

    # Predict
    df = df.copy()
    df['prob'] = rf.predict_proba(df[ind_var])[:, 1]
    df['prediction'] = rf.predict(df[ind_var])

    # Evaluation
    print(f'\n📊 {data_label.upper()} SET EVALUATION')
    print('\nConfusion Matrix:')
    print(confusion_matrix(df[dep_var], df['prediction']))

    accuracy = accuracy_score(df[dep_var], df['prediction'])
    precision = precision_score(df[dep_var], df['prediction'])
    recall = recall_score(df[dep_var], df['prediction'])
    f1 = f1_score(df[dep_var], df['prediction'])
    ll = log_loss(df[dep_var], df['prob'], labels=[0, 1])

    print(f'\nAccuracy: {accuracy:.2%}')
    print(f'Precision: {precision:.2%}')
    print(f'Recall: {recall:.2%}')
    print(f'F1 Score: {f1:.2%}')
    print(f'Log Loss: {ll:.2%}')

    # ROC Curve
    fpr, tpr, _ = roc_curve(df[dep_var], df['prob'])
    roc_auc = auc(fpr, tpr)
    plot_roc(fpr, tpr, roc_auc)

    # Feature Importance
    feature_importance = 100 * (rf.feature_importances_ / rf.feature_importances_.max())
    sorted_idx = np.argsort(feature_importance)[-10:]  # Top 10 only

    pos = np.arange(sorted_idx.shape[0]) + 0.5
    plt.figure()
    plt.barh(pos, feature_importance[sorted_idx], align='center', color='#28a745')
    plt.yticks(pos, np.array(ind_var)[sorted_idx], fontsize=8)
    plt.xlabel('Relative Feature Importance')
    plt.title(f'Random Forest Feature Importance ({data_label})')
    plt.tight_layout()
    plt.show()

    return {
        'accuracy': accuracy,
        'precision': precision,
        'recall': recall,
        'f1': f1,
        'log_loss': ll,
        'roc_auc': roc_auc,
        'top_features_idx': sorted_idx
    }


def run_neural_net_model(
    fix_train: pd.DataFrame,
    df: pd.DataFrame,
    ind_var: list[str],
    dep_var: str = 'team1_win',
    data_label: str = 'train',
    max_iter: int = 100,
    batch_size: int = 5,
    solver: str = 'sgd',
    show_feature_importance: bool = False,
) -> dict:
    """
    Trains and evaluates an MLP (Neural Network) classifier on a dataset.

    Parameters:
    ----------
    df : pd.DataFrame
        Dataset to run the model on.

    ind_var : list of str
        Independent variables.

    dep_var : str
        Dependent variable (default = 'team1_win').

    data_label : str
        Label for display (e.g., 'train', 'test').

    max_iter : int
        Max number of iterations.

    batch_size : int
        Batch size for MLP training.

    solver : str
        Solver to use for weight optimization (e.g., 'adam', 'sgd').

    show_feature_importance : bool
        Whether to calculate and display permutation feature importance.

    Returns:
    -------
    dict
        Model evaluation metrics and (optionally) top features.
    """

    # Initialize and train the model
    ann = MLP(max_iter=max_iter, batch_size=batch_size, solver=solver, random_state=42)
    ann.fit(fix_train[ind_var], fix_train[dep_var])

    # Predictions
    df = df.copy()
    df['prob'] = ann.predict_proba(df[ind_var])[:, 1]
    df['prediction'] = ann.predict(df[ind_var])

    # Metrics
    print(f'\n📊 {data_label.upper()} SET EVALUATION')
    print('\nConfusion Matrix:')
    print(confusion_matrix(df[dep_var], df['prediction']))

    accuracy = accuracy_score(df[dep_var], df['prediction'])
    precision = precision_score(df[dep_var], df['prediction'])
    recall = recall_score(df[dep_var], df['prediction'])
    f1 = f1_score(df[dep_var], df['prediction'])
    ll = log_loss(df[dep_var], df['prob'], labels=[0, 1])

    print(f'\nAccuracy: {accuracy:.2%}')
    print(f'Precision: {precision:.2%}')
    print(f'Recall: {recall:.2%}')
    print(f'F1 Score: {f1:.2%}')
    print(f'Log Loss: {ll:.2%}')

    # ROC & AUC
    fpr, tpr, _ = roc_curve(df[dep_var], df['prob'])
    roc_auc = auc(fpr, tpr)
    plot_roc(fpr, tpr, roc_auc)

    # Feature Importance (optional)
    if show_feature_importance:
        print("\n📌 Calculating permutation feature importance...")
        results = permutation_importance(ann, df[ind_var], df[dep_var], scoring='accuracy', random_state=42)
        feature_importance = 100 * (results.importances_mean / results.importances_mean.max())
        sorted_idx = np.argsort(feature_importance)[-10:]
        pos = np.arange(len(sorted_idx)) + 0.5

        plt.figure()
        plt.barh(pos, feature_importance[sorted_idx], align='center', color='#cc5500')
        plt.yticks(pos, np.array(ind_var)[sorted_idx], fontsize=8)
        plt.xlabel('Relative Feature Importance')
        plt.title(f'Neural Net Feature Importance ({data_label})')
        plt.tight_layout()
        plt.show()
    else:
        sorted_idx = None

    return {
        'accuracy': accuracy,
        'precision': precision,
        'recall': recall,
        'f1': f1,
        'log_loss': ll,
        'roc_auc': roc_auc,
        'top_features_idx': sorted_idx
    }


def run_xgboost_model(
    fix_train: pd.DataFrame,
    df: pd.DataFrame,
    ind_var: list[str],
    dep_var: str = 'team1_win',
    data_label: str = 'train',
    n_estimators: int = 100,
    learning_rate: float = 0.1,
    max_depth: int = 3,
    show_feature_importance: bool = True,
) -> dict:
    """
    Trains and evaluates an XGBoost classifier and plots metrics & feature importances.

    Parameters:
    ----------
    df : pd.DataFrame
        The input DataFrame (e.g., train or test set).

    ind_var : list of str
        List of independent variable names.

    dep_var : str
        Target variable name. Default is 'team1_win'.

    data_label : str
        Label to display in evaluation output (e.g., 'train' or 'test').

    n_estimators : int
        Number of boosting rounds.

    learning_rate : float
        Learning rate for gradient boosting.

    max_depth : int
        Maximum tree depth.

    show_feature_importance : bool
        If True, plots the top 10 most important features.

    Returns:
    -------
    dict
        Dictionary of performance metrics and top feature indices.
    """

    model = xgb.XGBClassifier(
        objective='binary:logistic',
        enable_categorical=True,
        n_estimators=n_estimators,
        learning_rate=learning_rate,
        max_depth=max_depth,
        use_label_encoder=False,
        eval_metric='logloss',
        random_state=42
    )

    model.fit(fix_train[ind_var], fix_train[dep_var])

    df = df.copy()
    df['prob'] = model.predict_proba(df[ind_var])[:, 1]
    df['prediction'] = model.predict(df[ind_var])

    print(f'\n📊 {data_label.upper()} SET EVALUATION')
    print('\nConfusion Matrix:')
    print(confusion_matrix(df[dep_var], df['prediction']))

    accuracy = accuracy_score(df[dep_var], df['prediction'])
    precision = precision_score(df[dep_var], df['prediction'])
    recall = recall_score(df[dep_var], df['prediction'])
    f1 = f1_score(df[dep_var], df['prediction'])
    ll = log_loss(df[dep_var], df['prob'], labels=[0, 1])

    print(f'\nAccuracy: {accuracy:.2%}')
    print(f'Precision: {precision:.2%}')
    print(f'Recall: {recall:.2%}')
    print(f'F1 Score: {f1:.2%}')
    print(f'Log Loss: {ll:.2%}')

    # ROC Curve
    fpr, tpr, _ = roc_curve(df[dep_var], df['prob'])
    roc_auc = auc(fpr, tpr)
    plot_roc(fpr, tpr, roc_auc)

    # Feature Importance Plot
    feature_importance = model.feature_importances_
    feature_importance = 100.0 * (feature_importance / feature_importance.max())
    sorted_idx = np.argsort(feature_importance)[-10:]  # Top 10

    if show_feature_importance:
        pos = np.arange(len(sorted_idx)) + 0.5
        plt.figure()
        plt.barh(pos, feature_importance[sorted_idx], align='center', color='#4e79a7')
        plt.yticks(pos, np.array(ind_var)[sorted_idx], fontsize=8)
        plt.xlabel('Relative Feature Importance')
        plt.title(f'XGBoost Feature Importance ({data_label})')
        plt.tight_layout()
        plt.show()

    return {
        'accuracy': accuracy,
        'precision': precision,
        'recall': recall,
        'f1': f1,
        'log_loss': ll,
        'roc_auc': roc_auc,
        'top_features_idx': sorted_idx
    }


# Correct function using literal_eval
def extract_ind_vars(df, col='Combination', save_to=None):
    ind_var = df[col].dropna().apply(ast.literal_eval).tolist()
    if save_to:
        with open(save_to, 'w') as f:
            f.write(str(ind_var))
    return ind_var


def run_model(
    fix_train: pd.DataFrame,
    ind_var_list,
    dataset: pd.DataFrame,
    target_col: str = 'team1_win',
    data_label: str = 'train',
    ):
    """
    Runs XGBoost, Random Forest, Logistic Regression, and Neural Network models
    over multiple feature combinations. Outputs results to an Excel file with
    sorted summaries and highlighted top 5 rows in each model tab.
    """

    # Map model names to corresponding model functions
    model_funcs = {
        'XGBoost': run_xgboost_model,
        'RandomForest': run_random_forest_model,
        'LogisticRegression': run_logistic_model,
        'NeuralNetwork': run_neural_net_model
    }

    # Create output directory
    output_dir = f'output/{data_label}_model_results'
    os.makedirs(output_dir, exist_ok=True)
    excel_path = os.path.join(output_dir, f'{data_label}_all_models_summary.xlsx')

    # Initialize Excel writer
    writer = pd.ExcelWriter(excel_path, engine='openpyxl')

    for model_name, model_func in model_funcs.items():
        summary = pd.DataFrame(columns=[
            'Model Serial', 'Model Variables', 'Accuracy', 'Precision', 'Recall', 'F Measure', 'AUC'
        ])

        for i, vars_ in enumerate(tqdm(ind_var_list, desc=f'Running {model_name}')):
            try:
                results = model_func(fix_train, dataset, ind_var=vars_, dep_var=target_col, data_label=data_label)
                summary.loc[i] = [
                    i,
                    vars_,
                    results['accuracy'],
                    results['precision'],
                    results['recall'],
                    results['f1'],
                    results['roc_auc']
                    ]
            except Exception as e:
                print(f"⚠️ Skipping model {i} for {model_name} due to error: {e}")

        # Sort by average and highlight top 5
        summary['Avg'] = summary[['Accuracy', 'Precision', 'Recall', 'F Measure', 'AUC']].mean(axis=1)
        summary_sorted = summary.sort_values(by='Avg', ascending=False).drop(columns='Avg')
        summary_sorted.reset_index(drop=True, inplace=True)

        # Save to sheet
        summary_sorted.to_excel(writer, index=False, sheet_name=model_name)

    writer.close()

    # Highlight top 5 rows in each sheet
    wb = load_workbook(excel_path)
    for sheet_name in model_funcs.keys():
        ws = wb[sheet_name]
        fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        for row in ws.iter_rows(min_row=2, max_row=6, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.fill = fill
    wb.save(excel_path)

    return excel_path

